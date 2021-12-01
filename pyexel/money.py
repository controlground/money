import tkinter
import tkinter.font
import tkinter.ttk
from openpyxl.reader.excel import load_workbook
import datetime

def 기록될_줄(sheet):
    row = 1
    while True:
        if sheet["A" + str(row)].value == None:
            break
        else:
            row += 1
    return row

def 입력은_숫자(money, entry_money):
    money = entry_money.get()
    if money.isdigit() == True:
        money = int(money)
        return money
    else:
        return False

def 확정():
    is_OK = True
    if 입력은_숫자(받은_돈, entry_받은_돈) == False:
        is_OK = False
    if 입력은_숫자(남은_돈, entry_남은_돈) == False:
        is_OK = False
    if 입력은_숫자(사용한_돈, entry_사용한_돈) == False:
        is_OK = False
    if is_OK == True:
        엑셀에_저장()
    else:
        경고.configure(text="입력이 유효하지 않습니다. 다시 입력값을 입력해주세요.")

def 엑셀에_저장():
    날짜 = str(datetime.datetime.now().date())
    내역 = entry_내역.get()
    row = 기록될_줄(sheet)
    if (row == 2 or sheet["F" + str((row - 1))].value == None):
        기존_전체_남은_돈 = 남은_돈 + 사용한_돈 - 받은_돈
    else:
        기존_전체_남은_돈 = int(sheet["F" + str((row - 1))].value)

    if 받은_돈 == 0:
        전체_남은_돈 = 기존_전체_남은_돈 - 사용한_돈
    else:
        전체_남은_돈 = 기존_전체_남은_돈 + 받은_돈
    row = str(row)
    sheet["A" + row] = 날짜
    sheet["B" + row] = 내역
    sheet["C" + row] = 받은_돈
    sheet["D" + row] = 사용한_돈
    sheet["E" + row] = 남은_돈
    sheet["F" + row] = 전체_남은_돈
    money.save("money.xlsx")
    초기화()
    경고.configure(text="")
    표_출력()

def 초기화():
    entry_내역.delete(0, "end")
    entry_받은_돈.delete(0, "end")
    entry_남은_돈.delete(0, "end")
    entry_사용한_돈.delete(0, "end")

def 삭제():
    #1. 삭제할 엑셀이 파이썬에 가져온다.
    money = load_workbook('money.xlsx')
    sheet = money["Sheet"]
    #2. 가장 최근 줄이 몇 째 줄인지 계산한다.
    row = 기록될_줄(sheet) - 1
    if row != 1:
        row = str(row)
        sheet["A" + row] = ""
        sheet["B" + row] = ""
        sheet["C" + row] = ""
        sheet["D" + row] = ""
        sheet["E" + row] = ""
        sheet["F" + row] = ""
        money.save("money.xlsx")
        표_출력()

def 표_출력():
    for i in 표.get_children():
        표.delete(i)
    row = 2
    while sheet['A' + str(row)].value != None:
        data = []
        data.append(sheet['B' + str(row)].value)
        data.append(sheet['C' + str(row)].value)
        data.append(sheet['D' + str(row)].value)
        data.append(sheet['E' + str(row)].value)
        data.append(sheet['F' + str(row)].value)
        표.insert('', 0, text=sheet["A"+str(row)].value, values=data)
        row += 1

window = tkinter.Tk()

window.title("mac")
window.geometry("1800x900")

money = load_workbook('money.xlsx')
sheet = money["Sheet"]

font = tkinter.font.Font(family="맑은 고딕", size=30)

내역 = tkinter.Label(window, text="내역", font=font)
내역.place(x=50, y=200)
entry_내역 = tkinter.Entry(window)
entry_내역.place(x=150, y=225)

받은_돈 = tkinter.Label(window, text="받은 돈", font=font)
받은_돈.place(x=50, y=300)
entry_받은_돈 = tkinter.Entry(window)
entry_받은_돈.place(x=200, y=325)
사용한_돈 = tkinter.Label(window, text="사용한 돈", font=font)
사용한_돈.place(x=50, y=400)
entry_사용한_돈 = tkinter.Entry(window)
entry_사용한_돈.place(x=250, y=425)
남은_돈 = tkinter.Label(window, text="남은 돈", font=font)
남은_돈.place(x=50, y=500)
entry_남은_돈 = tkinter.Entry(window)
entry_남은_돈.place(x=200, y=525)

btn_확정 = tkinter.Button(window, text="확정", width=7, height=3, command=확정)
btn_확정.place(x=50, y=600)
btn_초기화 = tkinter.Button(window, text="초기화", width=7, height=3, command=초기화)
btn_초기화.place(x=120, y=600)
btn_삭제 = tkinter.Button(window, text="삭제", width=7, height=3, command=삭제)
btn_삭제.place(x=190, y=600)

경고 = tkinter.Label(window, text="", font=font)
경고.place(x=50, y=50)

표 = tkinter.ttk.Treeview(window, columns=["A","B","C","D"], displaycolumns=["A","B","C","D"])
표.place(x=500, y=500)

표.column("#0", width=150)
표.heading("#0", text="날짜")

표.column("A", width=150)
표.heading("A", text="내역")

표.column("B", width=150)
표.heading("B", text="받은_돈")

표.column("C", width=150)
표.heading("C", text="사용한_돈")

표.column("D", width=150)
표.heading("D", text="남은_돈")

표_출력()


window.mainloop()


# 내역
# 받은 돈
# 사용한 돈
# 그날 남은 돈